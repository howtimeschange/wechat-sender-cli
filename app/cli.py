#!/usr/bin/env python3
"""
微信批量发送助手 — CLI 工具（macOS / Windows）
用法: python3 app/cli.py <命令>
"""
from __future__ import annotations

import argparse
import platform
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import openpyxl
import yaml
from rich.console import Console
from rich.table import Table
from rich.prompt import Confirm, Prompt
from rich.panel import Panel

console = Console()

IS_MAC = platform.system() == "Darwin"
IS_WINDOWS = platform.system() == "Windows"

ROOT = Path(__file__).resolve().parents[1]
CFG_PATH = ROOT / "config.yaml"
APPLE_SCRIPT = ROOT / "scripts" / "wechat_send_mac.applescript"
SHEET_TASKS = "发送任务"
HEADER_ROW = 2

STATUS_WAITING = "待发送"
STATUS_RUNNING = "发送中"
STATUS_SUCCESS = "发送成功"
STATUS_FAILED = "发送失败"

COL = {
    "seq": "#",
    "app": "* 应用",
    "target": "* 联系人/群聊",
    "msg_type": "* 消息类型",
    "text": "* 文字内容",
    "image": "图片路径",
    "send_time": "发送时间",
    "repeat": "重复",
    "remark": "备注",
    "status": "状态",
}


@dataclass
class Task:
    row: int
    app: str
    target: str
    msg_type: str
    text: str
    image_path: str
    send_time: Optional[datetime]
    repeat: str
    status: str


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def load_cfg() -> dict:
    if not CFG_PATH.exists():
        return {}
    with open(CFG_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def save_cfg(cfg: dict):
    with open(CFG_PATH, "w", encoding="utf-8") as f:
        yaml.dump(cfg, f, allow_unicode=True, default_flow_style=False)


def get_cfg():
    cfg = load_cfg()
    defaults = {
        "excel_path": "",
        "poll_seconds": 15,
        "dry_run": False,
        "send_interval": 5,
        "max_per_minute": 8,
    }
    for k, v in defaults.items():
        cfg.setdefault(k, v)
    return cfg


def find_columns(ws) -> dict:
    cols = {}
    for c in range(1, ws.max_column + 1):
        title = (ws.cell(HEADER_ROW, c).value or "").strip()
        if title:
            cols[title] = c
    return cols


def read_tasks(ws, cols) -> list[Task]:
    tasks: list[Task] = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        app = str(ws.cell(r, cols[COL["app"]]).value or "").strip()
        target = str(ws.cell(r, cols[COL["target"]]).value or "").strip()
        msg_type = str(ws.cell(r, cols[COL["msg_type"]]).value or "").strip()
        text = str(ws.cell(r, cols[COL["text"]]).value or "").strip()
        image_path = str(ws.cell(r, cols[COL["image"]]).value or "").strip()
        send_time_raw = ws.cell(r, cols[COL["send_time"]]).value
        repeat_raw = ws.cell(r, cols[COL["repeat"]]).value
        status = str(ws.cell(r, cols[COL["status"]]).value or "").strip()
        if not any([app, target, msg_type, text, image_path, send_time_raw, repeat_raw, status]):
            continue
        task = Task(
            row=r, app=app, target=target, msg_type=msg_type,
            text=text, image_path=image_path,
            send_time=send_time_raw if isinstance(send_time_raw, datetime) else None,
            repeat=str(repeat_raw).strip() if repeat_raw else "",
            status=status,
        )
        tasks.append(task)
    return tasks


def set_status(ws, cols, row: int, status: str):
    ws.cell(row, cols[COL["status"]]).value = status


def validate_task(task: Task):
    if task.app and task.app != "微信":
        raise ValueError(f"不支持的应用: {task.app}（当前仅支持微信）")
    if not task.target:
        raise ValueError("联系人/群聊不能为空")
    if task.msg_type not in {"文字", "图片", "文字+图片", "文字 "}:
        raise ValueError(f"不支持的消息类型: {task.msg_type}")
    if task.msg_type in {"文字", "文字+图片", "文字 "} and not task.text:
        raise ValueError("文字内容不能为空")
    if task.msg_type in {"图片", "文字+图片"}:
        if not task.image_path:
            raise ValueError("图片消息缺少图片路径")
        if not Path(task.image_path).expanduser().exists():
            raise ValueError(f"图片不存在: {task.image_path}")


def should_send(task: Task, now: datetime) -> bool:
    if task.status.startswith(STATUS_SUCCESS) and not task.repeat:
        return False
    if task.send_time is None:
        return True
    return now >= task.send_time


def call_sender(target: str, msg_type: str, text: str, image_path: str):
    if not IS_MAC:
        raise RuntimeError("微信自动化仅支持 macOS，请在 Mac 上运行")
    img = str(Path(image_path).expanduser()) if image_path else ""
    cmd = ["osascript", str(APPLE_SCRIPT), target, msg_type, text, img]
    r = subprocess.run(cmd, capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(r.stderr.strip() or "AppleScript 执行失败")


# ─── 命令 ────────────────────────────────────────────────

def cmd_config_show(_):
    cfg = get_cfg()
    os_label = "🍎 macOS" if IS_MAC else "🪟 Windows"
    table = Table(title=f"📋 当前配置 {os_label}", show_header=False, box=None)
    table.add_column("配置项", style="cyan")
    table.add_column("值", style="white")
    items = [
        ("表格路径 (excel_path)", cfg["excel_path"] or "(未设置)"),
        ("轮询间隔 (poll_seconds)", f"{cfg['poll_seconds']} 秒"),
        ("模拟运行 (dry_run)", "是 ✅" if cfg["dry_run"] else "否"),
        ("发送间隔 (send_interval)", f"{cfg['send_interval']} 秒"),
        ("每分钟最大条数 (max_per_minute)", str(cfg["max_per_minute"])),
        ("平台", "macOS（支持自动化）" if IS_MAC else "Windows（仅查看/配置）"),
    ]
    for k, v in items:
        table.add_row(k, v)
    console.print(table)


def cmd_config_set(args):
    cfg = get_cfg()
    key_map = {
        "excel_path": ("表格路径", str),
        "poll_seconds": ("轮询间隔", int),
        "dry_run": ("模拟运行", bool),
        "send_interval": ("发送间隔", float),
        "max_per_minute": ("每分钟最大条数", int),
    }
    if args.key not in key_map:
        console.print(f"[red]未知配置项: {args.key}[/red]")
        console.print("可选: " + ", ".join(key_map.keys()))
        return
    label, type_fn = key_map[args.key]
    try:
        value = type_fn(args.value)
    except ValueError as e:
        console.print(f"[red]格式错误: {e}[/red]")
        return
    cfg[args.key] = value
    save_cfg(cfg)
    console.print(f"[green]✅ 已更新 {label} = {value}[/green]")


def cmd_setup(_):
    console.print(Panel("🛠️  首次配置向导", expand=False))
    cfg = get_cfg()

    excel_path = Prompt.ask(
        "[cyan]表格路径[/cyan]",
        default=cfg.get("excel_path", "/Users/xingyicheng/Downloads/weme_batch_template.xlsx"),
    )
    send_interval = Prompt.ask(
        "[cyan]发送间隔（秒，两条消息间等待时间）[/cyan]",
        default=str(cfg.get("send_interval", 5)),
    )
    max_per_minute = Prompt.ask(
        "[cyan]每分钟最大发送条数[/cyan]",
        default=str(cfg.get("max_per_minute", 8)),
    )
    dry_run = Confirm.ask("[cyan]模拟运行（不真实发送）[/cyan]", default=cfg.get("dry_run", False))

    cfg["excel_path"] = excel_path
    cfg["send_interval"] = float(send_interval)
    cfg["max_per_minute"] = int(max_per_minute)
    cfg["dry_run"] = dry_run
    save_cfg(cfg)

    console.print("\n[green]✅ 配置已保存！[/green]")
    cmd_config_show(_)


def cmd_status(_):
    cfg = get_cfg()
    xlsx_path = Path(cfg.get("excel_path", "")).expanduser()
    if not xlsx_path.exists():
        console.print(f"[red]表格不存在: {xlsx_path}[/red]")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[SHEET_TASKS]
    cols = find_columns(ws)
    tasks = read_tasks(ws, cols)

    waiting = sum(1 for t in tasks if t.status in ("", STATUS_WAITING) or not t.status)
    running = sum(1 for t in tasks if t.status == STATUS_RUNNING)
    success = sum(1 for t in tasks if t.status.startswith(STATUS_SUCCESS))
    failed = sum(1 for t in tasks if t.status.startswith(STATUS_FAILED))

    table = Table(title=f"📊 任务概览 — {xlsx_path.name.replace(chr(10), '')}", box=None)
    table.add_column("状态", style="white")
    table.add_column("数量", justify="right")
    table.add_row("⏳ 待发送", f"[yellow]{waiting}[/yellow]")
    table.add_row("🔄 发送中", f"[blue]{running}[/blue]")
    table.add_row("✅ 发送成功", f"[green]{success}[/green]")
    table.add_row("❌ 发送失败", f"[red]{failed}[/red]")
    table.add_row("📄 总计", str(len(tasks)))
    console.print(table)

    if not IS_MAC:
        console.print("\n[yellow]⚠️  当前为 Windows 环境，无法执行自动化发送，仅支持查看状态[/yellow]")

    pending = [t for t in tasks if not t.status.startswith(STATUS_SUCCESS)]
    if pending:
        task_table = Table(title="📋 待发送任务", box=None)
        task_table.add_column("#", justify="right", style="dim")
        task_table.add_column("类型")
        task_table.add_column("目标")
        task_table.add_column("内容预览")
        task_table.add_column("状态")
        for i, t in enumerate(pending[:20], 1):
            preview = (t.text or t.image_path or "")[:30]
            status_str = t.status or STATUS_WAITING
            task_table.add_row(str(i), t.msg_type or "-", t.target[:15], preview, status_str)
        console.print(task_table)
        if len(pending) > 20:
            console.print(f"[dim]...还有 {len(pending) - 20} 条[/dim]")


def cmd_send(args):
    if not IS_MAC:
        console.print("[red]❌ 微信自动化发送仅支持 macOS[/red]")
        console.print("[yellow]请在 Mac 上运行此命令[/yellow]")
        return

    cfg = get_cfg()
    xlsx_path = Path(cfg.get("excel_path", "")).expanduser()
    if not xlsx_path.exists():
        console.print(f"[red]表格不存在: {xlsx_path}[/red]")
        return

    dry_run = cfg.get("dry_run", False)
    send_interval = cfg.get("send_interval", 5)
    max_per_minute = cfg.get("max_per_minute", 8)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[SHEET_TASKS]
    cols = find_columns(ws)
    tasks = read_tasks(ws, cols)
    now = datetime.now()

    pending = [t for t in tasks if should_send(t, now)]
    if not pending:
        console.print("[yellow]没有需要发送的任务[/yellow]")
        return

    console.print(f"[cyan]开始发送 {len(pending)} 条任务...[/cyan]")
    if dry_run:
        console.print("[yellow]⚠️  模拟运行模式，不会真实发送[/yellow]")

    sent_times: list[datetime] = []
    success_count, fail_count = 0, 0

    for i, task in enumerate(pending):
        window_start = now - timedelta(minutes=1)
        sent_times = [t for t in sent_times if t > window_start]
        if len(sent_times) >= max_per_minute:
            sleep_secs = 60 - (datetime.now() - sent_times[0]).total_seconds()
            if sleep_secs > 0:
                console.print(f"[dim]达到每分钟 {max_per_minute} 条限制，等待 {sleep_secs:.0f}s...[/dim]")
                time.sleep(sleep_secs)
                sent_times = [t for t in sent_times if t > datetime.now() - timedelta(minutes=1)]

        console.print(f"[{i+1}/{len(pending)}] → {task.target} [{task.msg_type}]", end="")

        set_status(ws, cols, task.row, STATUS_RUNNING)
        wb.save(xlsx_path)

        try:
            if not dry_run:
                validate_task(task)
                call_sender(task.target, task.msg_type, task.text, task.image_path)
            status = f"{STATUS_SUCCESS} {now.strftime('%H:%M:%S')}"
            set_status(ws, cols, task.row, status)
            sent_times.append(datetime.now())
            success_count += 1
            console.print(f" [green]✅[/green]")
        except Exception as e:
            status = f"{STATUS_FAILED}: {e}"
            set_status(ws, cols, task.row, status)
            fail_count += 1
            console.print(f" [red]❌ {e}[/red]")

        wb.save(xlsx_path)

        if i < len(pending) - 1 and not dry_run:
            time.sleep(send_interval)

    console.print(f"\n[green]完成！成功 {success_count} 条，失败 {fail_count} 条[/green]")


def cmd_daemon(args):
    if not IS_MAC:
        console.print("[red]❌ 微信自动化发送仅支持 macOS[/red]")
        return

    cfg = get_cfg()
    xlsx_path = Path(cfg.get("excel_path", "")).expanduser()
    poll_seconds = int(cfg.get("poll_seconds", 15))
    dry_run = cfg.get("dry_run", False)
    send_interval = cfg.get("send_interval", 5)
    max_per_minute = cfg.get("max_per_minute", 8)

    if not xlsx_path.exists():
        console.print(f"[red]表格不存在: {xlsx_path}[/red]")
        return

    console.print(f"[cyan]守护进程模式启动 | 间隔 {poll_seconds}s | dry_run={dry_run}[/cyan]")
    sent_times: list[datetime] = []

    while True:
        try:
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb[SHEET_TASKS]
            cols = find_columns(ws)
            tasks = read_tasks(ws, cols)
            now = datetime.now()
            changed = False

            for task in tasks:
                if task.status.startswith(STATUS_SUCCESS) and not task.repeat:
                    continue
                try:
                    validate_task(task)
                except Exception as e:
                    set_status(ws, cols, task.row, f"{STATUS_FAILED}: {e}")
                    changed = True
                    continue
                if not should_send(task, now):
                    if not task.status:
                        set_status(ws, cols, task.row, STATUS_WAITING)
                        changed = True
                    continue

                window_start = now - timedelta(minutes=1)
                sent_times = [t for t in sent_times if t > window_start]
                if len(sent_times) >= max_per_minute:
                    time.sleep(1)
                    continue

                set_status(ws, cols, task.row, STATUS_RUNNING)
                changed = True

                try:
                    if not dry_run:
                        call_sender(task.target, task.msg_type, task.text, task.image_path)
                    status = f"{STATUS_SUCCESS} {now.strftime('%m-%d %H:%M')}"
                    set_status(ws, cols, task.row, status)
                    sent_times.append(datetime.now())
                    time.sleep(send_interval)
                except Exception as e:
                    set_status(ws, cols, task.row, f"{STATUS_FAILED}: {e}")
                    changed = True

            if changed:
                wb.save(xlsx_path)

        except Exception as e:
            console.print(f"[red]错误: {e}[/red]")

        time.sleep(poll_seconds)


def cmd_template(_):
    console.print("[cyan]请使用 Excel 编辑模板文件[/cyan]")
    console.print("\n列名如下（必须与模板一致）：")
    table = Table(box=None)
    table.add_column("列名", style="yellow")
    table.add_column("说明", style="white")
    table.add_column("示例", style="dim")
    rows = [
        ("* 应用", "固定填 微信", "微信"),
        ("* 联系人/群聊", "对方微信名或群名", "张三"),
        ("* 消息类型", "文字 / 图片 / 文字+图片", "文字"),
        ("* 文字内容", "要发送的文本", "你好！"),
        ("图片路径", "本地图片绝对路径（macOS）", "/Users/xx/pic.png"),
        ("发送时间", "格式 2025-01-01 14:30", "2025-01-01 14:30"),
        ("重复", "daily / weekly / workday / 空", "daily"),
        ("状态", "自动填充，勿手动编辑", "发送成功"),
    ]
    for r in rows:
        table.add_row(*r)
    console.print(table)


def main():
    parser = argparse.ArgumentParser(
        description="微信批量发送助手 CLI (macOS / Windows)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sub = parser.add_subparsers(dest="cmd")

    p_config = sub.add_parser("config", help="查看 / 修改配置")
    p_config.add_argument("key", nargs="?", help="配置项名称")
    p_config.add_argument("value", nargs="?", help="新的配置值")
    p_config.set_defaults(func=lambda a: cmd_config_set(a) if a.key else cmd_config_show(None))

    sub.add_parser("setup", help="交互式配置向导").set_defaults(func=cmd_setup)
    sub.add_parser("status", help="查看任务状态").set_defaults(func=cmd_status)
    sub.add_parser("send", help="立即发送待处理任务（仅 macOS）").set_defaults(func=cmd_send)
    sub.add_parser("daemon", help="守护进程模式（仅 macOS）").set_defaults(func=cmd_daemon)
    sub.add_parser("template", help="查看表格模板格式").set_defaults(func=cmd_template)

    args = parser.parse_args()

    if args.cmd is None:
        parser.print_help()
        console.print("\n[dim]快速上手:[/dim]  python3 app/cli.py setup")
        return

    args.func(args)


if __name__ == "__main__":
    main()
