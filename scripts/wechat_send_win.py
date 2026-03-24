#!/usr/bin/env python3
"""
微信批量发送助手 — Windows 版批量发送核心
依赖: pip install uiautomation pyperclip pywin32 Pillow openpyxl PyYAML rich
用法: python scripts/wechat_send_win.py [--dry]
"""
from __future__ import annotations

import argparse
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

# 自动安装缺失的 Windows 自动化依赖
try:
    import uiautomation as auto
    import pyperclip
    import win32clipboard
    import win32con
    from PIL import Image
except ImportError:
    print("[信息] 正在安装 Windows 自动化依赖...")
    subprocess.run([sys.executable, "-m", "pip", "install",
                   "uiautomation", "pyperclip", "pywin32", "Pillow", "-q"])
    import uiautomation as auto
    import pyperclip
    import win32clipboard
    import win32con
    from PIL import Image

# ─── 配置（从外部 config.yaml 读取） ──────────────────────

ROOT = Path(__file__).resolve().parents[1]
CFG_PATH = ROOT / "config.yaml"
XLSX_PATH = None  # 从 config 读取

SHEET_TASKS = "发送任务"
HEADER_ROW = 2

STATUS_WAITING = "待发送"
STATUS_RUNNING = "发送中"
STATUS_SUCCESS = "发送成功"
STATUS_FAILED = "发送失败"

COL = {
    "seq": "#",
    "app": "* 应用",
    "target": "* 联系人/群聊",
    "msg_type": "* 消息类型",
    "text": "* 文字内容",
    "image": "图片路径",
    "send_time": "发送时间",
    "repeat": "重复",
    "remark": "备注",
    "status": "状态",
}


def load_cfg() -> dict:
    if not CFG_PATH.exists():
        return {}
    import yaml
    with open(CFG_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


@dataclass
class Task:
    row: int
    app: str
    target: str
    msg_type: str
    text: str
    image_path: str
    send_time: Optional[datetime]
    repeat: str
    status: str


def find_columns(ws):
    cols = {}
    for c in range(1, ws.max_column + 1):
        title = (ws.cell(HEADER_ROW, c).value or "").strip()
        if title:
            cols[title] = c
    missing = [v for k, v in COL.items() if v not in cols and not k.startswith("seq")]
    if missing:
        raise RuntimeError(f"模板缺少列: {missing}")
    return cols


def read_tasks(ws, cols) -> list[Task]:
    tasks = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        app = str(ws.cell(r, cols[COL["app"]]).value or "").strip()
        target = str(ws.cell(r, cols[COL["target"]]).value or "").strip()
        msg_type = str(ws.cell(r, cols[COL["msg_type"]]).value or "").strip()
        text = str(ws.cell(r, cols[COL["text"]]).value or "").strip()
        image_path = str(ws.cell(r, cols[COL["image"]]).value or "").strip()
        send_time_raw = ws.cell(r, cols[COL["send_time"]]).value
        repeat_raw = ws.cell(r, cols[COL["repeat"]]).value
        status = str(ws.cell(r, cols[COL["status"]]).value or "").strip()
        if not any([app, target, msg_type, text, image_path, send_time_raw, repeat_raw, status]):
            continue
        send_time = send_time_raw if isinstance(send_time_raw, datetime) else None
        tasks.append(Task(
            row=r, app=app, target=target, msg_type=msg_type,
            text=text, image_path=image_path,
            send_time=send_time,
            repeat=str(repeat_raw).strip() if repeat_raw else "",
            status=status,
        ))
    return tasks


def set_status(ws, cols, row: int, status: str):
    ws.cell(row, cols[COL["status"]]).value = status


def should_send(task: Task, now: datetime) -> bool:
    if task.status.startswith(STATUS_SUCCESS) and not task.repeat:
        return False
    if task.send_time is None:
        return True
    return now >= task.send_time


# ─── Windows 微信自动化 ─────────────────────────────────

WECHAT_WIN_TITLE = "微信"


def check_wechat() -> bool:
    win = auto.WindowControl(Name=WECHAT_WIN_TITLE, searchDepth=1)
    return win.Exists(0)


def activate_wechat():
    win = auto.WindowControl(Name=WECHAT_WIN_TITLE, searchDepth=1)
    if not win.Exists(3):
        raise RuntimeError("未找到微信窗口，请先打开并登录 PC 微信")
    win.SetActive()
    time.sleep(0.4)
    return win


def search_contact(name: str):
    auto.SendKeys("{Ctrl}f")
    time.sleep(0.25)
    pyperclip.copy(name)
    auto.SendKeys("{Ctrl}v")
    time.sleep(0.2)
    auto.SendKeys("{Enter}")
    time.sleep(0.5)


def send_text(text: str):
    pyperclip.copy(text)
    auto.SendKeys("{Ctrl}v")
    time.sleep(0.15)
    auto.SendKeys("{Enter}")


def send_image(image_path: str):
    img = Image.open(image_path).convert("RGB")
    win32clipboard.OpenClipboard()
    try:
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardData(win32con.CF_BITMAP, img)
    finally:
        win32clipboard.CloseClipboard()
    time.sleep(0.1)
    auto.SendKeys("{Ctrl}v")
    time.sleep(0.2)
    auto.SendKeys("{Enter}")


def send_text_with_image(text: str, image_path: str):
    pyperclip.copy(text)
    auto.SendKeys("{Ctrl}v")
    time.sleep(0.15)
    send_image(image_path)


def call_send(target: str, msg_type: str, text: str, image_path: str):
    activate_wechat()
    search_contact(target)
    if msg_type == "文字":
        send_text(text)
    elif msg_type == "图片":
        send_image(image_path)
    elif msg_type == "文字+图片":
        send_text_with_image(text, image_path)
    else:
        raise ValueError(f"不支持的消息类型: {msg_type}")


# ─── 主发送逻辑 ─────────────────────────────────────────

def batch_send(dry_run: bool = False, send_interval: float = 5,
               max_per_minute: int = 8, xlsx_path: str = ""):
    import openpyxl

    if not xlsx_path:
        raise ValueError("未设置 excel_path，请先运行 python app/cli.py setup 配置表格路径")
    xlsx_path = Path(xlsx_path).expanduser()
    if not xlsx_path.exists():
        raise RuntimeError(f"表格不存在: {xlsx_path}")

    print(f"📋 读取表格: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[SHEET_TASKS]
    cols = find_columns(ws)
    tasks = read_tasks(ws, cols)
    now = datetime.now()

    pending = [t for t in tasks if should_send(t, now)]
    if not pending:
        print("⏳ 没有需要发送的任务")
        return

    print(f"📤 开始发送 {len(pending)} 条任务...")
    if dry_run:
        print("⚠️  模拟运行模式，不会真实发送")

    sent_times: list[datetime] = []
    success_count, fail_count = 0, 0

    for i, task in enumerate(pending):
        # 频率控制
        window_start = now - timedelta(minutes=1)
        sent_times = [t for t in sent_times if t > window_start]
        if len(sent_times) >= max_per_minute:
            sleep_secs = 60 - (datetime.now() - sent_times[0]).total_seconds()
            if sleep_secs > 0:
                print(f"⏳ 达到每分钟 {max_per_minute} 条限制，等待 {sleep_secs:.0f}s...")
                time.sleep(sleep_secs)
                sent_times = [t for t in sent_times if t > datetime.now() - timedelta(minutes=1)]

        print(f"[{i+1}/{len(pending)}] → {task.target} [{task.msg_type}]", end="")

        set_status(ws, cols, task.row, STATUS_RUNNING)
        wb.save(xlsx_path)

        try:
            if not dry_run:
                if task.app and task.app != "微信":
                    raise ValueError(f"不支持的应用: {task.app}（当前仅支持微信）")
                if not task.target:
                    raise ValueError("联系人/群聊不能为空")
                if task.msg_type not in {"文字", "图片", "文字+图片"}:
                    raise ValueError(f"不支持的消息类型: {task.msg_type}")
                if task.msg_type in {"文字", "文字+图片"} and not task.text:
                    raise ValueError("文字内容不能为空")
                call_send(task.target, task.msg_type, task.text, task.image_path)

            status = f"{STATUS_SUCCESS} {datetime.now().strftime('%H:%M:%S')}"
            set_status(ws, cols, task.row, status)
            sent_times.append(datetime.now())
            success_count += 1
            print(" ✅")
        except Exception as e:
            set_status(ws, cols, task.row, f"{STATUS_FAILED}: {e}")
            fail_count += 1
            print(f" ❌ {e}")

        wb.save(xlsx_path)

        if i < len(pending) - 1 and not dry_run:
            time.sleep(send_interval)

    print(f"\n✅ 完成！成功 {success_count} 条，失败 {fail_count} 条")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="微信批量发送 — Windows 版")
    parser.add_argument("--dry", action="store_true", help="模拟运行")
    args = parser.parse_args()

    cfg = load_cfg()
    xlsx_path = cfg.get("excel_path", "")
    send_interval = float(cfg.get("send_interval", 5))
    max_per_minute = int(cfg.get("max_per_minute", 8))
    dry_run = args.dry or cfg.get("dry_run", False)

    batch_send(
        dry_run=dry_run,
        send_interval=send_interval,
        max_per_minute=max_per_minute,
        xlsx_path=xlsx_path,
    )
